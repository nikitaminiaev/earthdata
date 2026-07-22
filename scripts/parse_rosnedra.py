import openpyxl
import json
import re
import sys
from pathlib import Path

VECTORS_DIR = Path(__file__).resolve().parent.parent / "data" / "vectors"
XLSX_PATH = "/tmp/rosnedra_2026.xlsx"
OUTPUT = VECTORS_DIR / "rosnedra_plots.geojson"

# Region centroids (fallback when coordinates parsing fails)
REGION_COORDS = {
    "Самарская область": (53.2, 50.5),
    "Пермский край": (59.0, 56.5),
    "Республика Саха (Якутия)": (65.0, 126.0),
    "Ханты-Мансийский автономный округ-Югра": (61.5, 70.0),
    "Удмуртская Республика": (57.0, 53.0),
    "Чувашская Республика": (55.5, 47.0),
    "Ненецкий автономный округ": (68.0, 55.0),
    "Саратовская область": (51.5, 47.0),
    "Иркутская область": (55.0, 105.0),
    "Республика Башкортостан": (54.5, 56.0),
    "Астраханская область": (47.0, 48.0),
    "Ямало-Ненецкий автономный округ": (67.0, 75.0),
    "Красноярский край": (65.0, 95.0),
    "Республика Коми": (64.0, 55.0),
    "Республика Татарстан": (55.5, 50.5),
    "Оренбургская область": (52.0, 55.0),
    "Томская область": (58.5, 80.0),
    "Кемеровская область": (55.0, 87.0),
    "Республика Бурятия": (53.0, 109.0),
    "Забайкальский край": (52.0, 115.0),
    "Мурманская область": (68.0, 35.0),
    "Республика Карелия": (63.0, 34.0),
    "Ленинградская область": (60.0, 32.0),
    "Вологодская область": (60.0, 40.0),
    "Архангельская область": (64.0, 42.0),
    "Кировская область": (58.0, 50.0),
    "Нижегородская область": (56.0, 44.0),
    "Ульяновская область": (54.0, 48.0),
    "Пензенская область": (53.0, 45.0),
    "Тамбовская область": (52.5, 41.5),
    "Липецкая область": (52.5, 39.0),
    "Белгородская область": (50.5, 36.5),
    "Курская область": (51.5, 36.0),
    "Воронежская область": (51.0, 40.0),
    "Волгоградская область": (49.5, 44.0),
    "Ростовская область": (47.5, 41.0),
    "Краснодарский край": (45.5, 39.5),
    "Ставропольский край": (45.0, 43.0),
    "Республика Крым": (45.0, 34.5),
    "Свердловская область": (57.5, 62.0),
    "Челябинская область": (55.0, 61.0),
    "Тюменская область": (58.0, 72.0),
    "Омская область": (55.0, 73.0),
    "Новосибирская область": (55.0, 82.0),
    "Алтайский край": (52.5, 83.0),
    "Республика Алтай": (50.5, 87.0),
    "Хабаровский край": (52.0, 136.0),
    "Приморский край": (45.0, 134.0),
    "Амурская область": (52.0, 128.0),
    "Еврейская автономная область": (48.5, 132.5),
    "Магаданская область": (62.0, 150.0),
    "Камчатский край": (56.0, 160.0),
    "Сахалинская область": (50.0, 143.0),
    "Чукотский автономный округ": (67.0, 175.0),
}

MINERAL_COLORS = {
    "углеводородное сырье": ("#e74c3c", "Нефть/газ"),
    "нефть": ("#e74c3c", "Нефть"),
    "газ": ("#c0392b", "Газ"),
    "воды подземные": ("#3498db", "Подземные воды"),
    "воды подземные промышленные": ("#2980b9", "Пром. воды"),
    "твердые полезные ископаемые": ("#e67e22", "ТПИ"),
    "тпи": ("#e67e22", "ТПИ"),
    "золото": ("#f1c40f", "Золото"),
    "руда": ("#d35400", "Руда"),
}


def _compute_centroid(points_gsk, points_sk):
    centroid = None
    if points_gsk:
        lats = [p[0] for p in points_gsk if p[0] is not None]
        lons = [p[1] for p in points_gsk if p[1] is not None]
        if lats and lons:
            centroid = (sum(lats) / len(lats), sum(lons) / len(lons))
    if centroid is None and points_sk:
        lats = [p[0] for p in points_sk if p[0] is not None]
        lons = [p[1] for p in points_sk if p[1] is not None]
        if lats and lons:
            centroid = (sum(lats) / len(lats), sum(lons) / len(lons))
    return centroid


def _region_to_coords(region_name):
    for reg_key, coord in REGION_COORDS.items():
        if reg_key.lower() in region_name.lower() or region_name.lower() in reg_key.lower():
            return coord
    return REGION_COORDS.get(region_name, None)


def _make_feature(pid, mineral_type, plot_name, region, area, reserves,
                  usage_type, auction_form, auction_timing, sheet_name, centroid):
    mineral_cat = extract_mineral_category(mineral_type) if mineral_type else "другое"
    color_info = MINERAL_COLORS.get(mineral_cat, ("#95a5a6", mineral_cat.title()))

    def _clean(s):
        if not s:
            return ""
        return " ".join(s.replace("\n", " ").split())

    return {
        "type": "Feature",
        "geometry": {
            "type": "Point",
            "coordinates": [centroid[1], centroid[0]]
        },
        "properties": {
            "id": pid,
            "mineral": _clean(mineral_type),
            "mineral_cat": mineral_cat,
            "name": _clean(plot_name),
            "region": _clean(region),
            "area": _clean(str(area)) if area else "",
            "reserves": _clean(str(reserves))[:300] if reserves else "",
            "usage": _clean(usage_type),
            "form": _clean(auction_form),
            "timing": _clean(auction_timing),
            "source": sheet_name,
            "marker-color": color_info[0],
            "marker-size": "small",
        },
    }


def parse_latlon(val_deg, val_min, val_sec):
    try:
        d = float(str(val_deg).replace(",", "."))
        m = float(str(val_min).replace(",", "."))
        s = float(str(val_sec).replace(",", ".").replace(" ", ""))
        return d + m / 60 + s / 3600
    except (ValueError, TypeError):
        return None


def extract_mineral_category(mineral_str):
    if not mineral_str:
        return "другое"
    ms = mineral_str.lower().strip()
    if "углеводород" in ms:
        return "углеводородное сырье"
    if "подземн" in ms and "вод" in ms:
        return "подземные воды"
    if "тверд" in ms:
        return "твердые полезные ископаемые"
    if "золот" in ms:
        return "золото"
    if "желез" in ms:
        return "руда"
    if "полиметал" in ms:
        return "руда"
    if "мед" in ms:
        return "руда"
    if "марганц" in ms:
        return "руда"
    if "алмаз" in ms:
        return "алмазы"
    if "уголь" in ms:
        return "уголь"
    return "другое"


def dms_to_dd(d, m, s):
    if d is None or m is None or s is None:
        return None
    try:
        d = float(str(d).replace(",", ".").strip())
        m = float(str(m).replace(",", ".").strip())
        s = float(str(s).replace(",", ".").replace(" ", "").strip())
        return d + m / 60 + s / 3600
    except (ValueError, TypeError):
        return None


def parse_rosnedra():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    features = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find first data row: has numeric col 0 >= 1 AND mineral type in col 1/2
        data_start = None
        current_region = None

        for i, row in enumerate(rows):
            if i < 8:
                continue
            col0 = row[0]
            col1 = row[1] if len(row) > 1 else None
            col2 = str(row[2]) if len(row) > 2 and row[2] else ""

            # Column-header number rows (1, 2, 3, 4) have numbers only in early cols
            # Skip them
            is_col_header = col0 is not None and col1 is None and col2 == ""
            if is_col_header and isinstance(col0, (int, float)) and col0 in (1, 2, 3, 4, '1', '2', '3'):
                continue

            # Check for data row: numeric col0 >= 1 AND col1 has mineral keyword
            if col0 is not None:
                pid = None
                if isinstance(col0, (int, float)):
                    pid = int(col0)
                elif isinstance(col0, str):
                    try:
                        pid = int(col0.strip())
                    except ValueError:
                        pass
                if pid is not None and pid >= 1:
                    # Check if col1 or col2 has mineral content
                    mineral_hint = str(col1 or "") + str(col2 or "")
                    if any(kw in mineral_hint.lower() for kw in
                           ["сырье", "нефт", "газ", "вод", "полезн", "золот", "руд", "уголь"]):
                        data_start = i
                        # Scan backwards for region name
                        for j in range(data_start - 1, max(data_start - 5, 0), -1):
                            prev = rows[j]
                            pv = str(prev[0]) if prev[0] else ""
                            if prev[0] and isinstance(prev[0], str) and \
                               len(pv) > 5 and not pv.startswith("№") and \
                               not pv.startswith("1") and not any(kw in pv for kw in
                                ["с.ш", "в.д", "град", "мин", "сек", "п/п"]):
                                current_region = pv.strip()
                                break
                        break

        if data_start is None:
            continue

        # Group rows by plot
        plot_id = 0
        ec = len(rows[0]) if rows else 0
        sheet_features = []
        has_plot = False
        mineral_type = None
        plot_name = None
        plot_area = None
        reserves = None
        protocol = None
        usage_type = None
        auction_form = None
        auction_timing = None
        points_gsk = []
        points_sk = []

        region_updated_this_loop = False

        for i in range(data_start, len(rows)):
            row = rows[i]
            first = row[0] if len(row) > 0 else None
            second = row[1] if len(row) > 1 else None
            region_updated_this_loop = False

            # Check if this is a region marker (region name in col 0, no data in cols 1-3)
            if first is not None and isinstance(first, str) and not first[0].isdigit():
                # Skip column header rows
                if first.strip() in ("№", "№ п/п", "№ П/П", "1", "2"):
                    continue
                col1_empty = row[1] is None if len(row) > 1 else True
                col2_empty = row[2] is None if len(row) > 2 else True
                if col1_empty and col2_empty and len(first.strip()) > 3:
                    current_region = first.strip()
                    plot_id = 0  # reset per-region numbering
                    region_updated_this_loop = True
                    continue

            # Check if this is a new plot (has ID in column A AND plot name in col 2)
            is_new_plot = False
            if first is not None:
                col2_val = row[2] if len(row) > 2 else None
                if isinstance(first, (int, float)):
                    if first >= 1 and col2_val is not None:
                        is_new_plot = True
                elif isinstance(first, str):
                    stripped = first.strip()
                    try:
                        num_val = int(stripped)
                        if num_val >= 1 and col2_val is not None:
                            is_new_plot = True
                    except ValueError:
                        pass

            if is_new_plot:
                if has_plot:
                    centroid = _compute_centroid(points_gsk, points_sk)
                    if centroid is None and current_region:
                        centroid = _region_to_coords(current_region)

                    if centroid:
                        feature = _make_feature(plot_id, mineral_type, plot_name,
                                               current_region, plot_area, reserves,
                                               usage_type, auction_form, auction_timing,
                                               sheet_name, centroid)
                        sheet_features.append(feature)

                has_plot = True
                plot_id = int(first)
                mineral_type = str(row[1]).strip() if len(row) > 1 and row[1] else None
                plot_name = str(row[2]).strip() if len(row) > 2 and row[2] else None
                plot_area = str(row[3]).strip() if len(row) > 3 and row[3] else None
                # Reserves are in column 17 or 18 depending on sheet
                reserves = str(row[17]).strip() if len(row) > 17 and row[17] else None
                protocol = str(row[18]).strip() if len(row) > 18 and row[18] else None
                usage_type = str(row[19]).strip() if len(row) > 19 and row[19] else None
                auction_form = str(row[20]).strip() if len(row) > 20 and row[20] else None
                auction_timing = str(row[21]).strip() if len(row) > 21 and row[21] else None
                points_gsk = []
                points_sk = []

                # Parse first point coordinates
                if len(row) > 8:
                    lat = dms_to_dd(row[5] if len(row) > 5 else None,
                                    row[6] if len(row) > 6 else None,
                                    row[7] if len(row) > 7 else None)
                    lon = dms_to_dd(row[8] if len(row) > 8 else None,
                                    row[9] if len(row) > 9 else None,
                                    row[10] if len(row) > 10 else None)
                    if lat is not None and lon is not None:
                        points_gsk.append((lat, lon))
                if len(row) > 14:
                    lat_sk = dms_to_dd(row[11] if len(row) > 11 else None,
                                       row[12] if len(row) > 12 else None,
                                       row[13] if len(row) > 13 else None)
                    lon_sk = dms_to_dd(row[14] if len(row) > 14 else None,
                                       row[15] if len(row) > 15 else None,
                                       row[16] if len(row) > 16 else None)
                    if lat_sk is not None and lon_sk is not None:
                        points_sk.append((lat_sk, lon_sk))

            else:
                # Continuation of current plot - add more coordinate points
                if len(row) > 8:
                    lat = dms_to_dd(row[5] if len(row) > 5 else None,
                                    row[6] if len(row) > 6 else None,
                                    row[7] if len(row) > 7 else None)
                    lon = dms_to_dd(row[8] if len(row) > 8 else None,
                                    row[9] if len(row) > 9 else None,
                                    row[10] if len(row) > 10 else None)
                    if lat is not None and lon is not None:
                        points_gsk.append((lat, lon))
                if len(row) > 14:
                    lat_sk = dms_to_dd(row[11] if len(row) > 11 else None,
                                       row[12] if len(row) > 12 else None,
                                       row[13] if len(row) > 13 else None)
                    lon_sk = dms_to_dd(row[14] if len(row) > 14 else None,
                                       row[15] if len(row) > 15 else None,
                                       row[16] if len(row) > 16 else None)
                    if lat_sk is not None and lon_sk is not None:
                        points_sk.append((lat_sk, lon_sk))

        # Don't forget the last plot
        if has_plot:
            centroid = _compute_centroid(points_gsk, points_sk)
            if centroid is None and current_region:
                centroid = _region_to_coords(current_region)

            if centroid:
                feature = _make_feature(plot_id, mineral_type, plot_name,
                                       current_region, plot_area, reserves,
                                       usage_type, auction_form, auction_timing,
                                       sheet_name, centroid)
                sheet_features.append(feature)

        features.extend(sheet_features)
        print(f"  Sheet '{sheet_name}': parsed {len(sheet_features)} plots, region={current_region}")

    # Also download and parse the main base sheet from the 26.01.2026 entry
    # (it's the same XLSX but might have additional sheets not in our file)

    fc = {
        "type": "FeatureCollection",
        "features": features,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(fc, f, ensure_ascii=False, indent=2)

    print(f"\nTotal: {len(features)} plots saved to {OUTPUT}")


def main():
    if not Path(XLSX_PATH).exists():
        print(f"XLSX not found at {XLSX_PATH}")
        print("Download from: https://rosnedra.gov.ru/upload/iblock/eb3/s4yur7z3izbryz7zc5a5jtft9r0ilx8w.xlsx")
        sys.exit(1)
    parse_rosnedra()


if __name__ == "__main__":
    main()
